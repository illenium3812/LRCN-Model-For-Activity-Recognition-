from tensorflow_docs.vis import embed
from tensorflow import keras
from imutils import paths

import matplotlib.pyplot as plt
import tensorflow as tf
import pandas as pd
import numpy as np
import imageio
import cv2
import os
import beepy
import pandas as pd
import json
import time
import datetime
import psycopg2
from openpyxl import load_workbook


import datetime
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import drawing
import cv2
import win32com.client


import cv2
import numpy as np
from collections import deque
import PIL
from tensorflow import keras


print("import done")

def lst2pgarr(alist):
    return '{"' + '","'.join(alist) + '"}'


def get_employee_header():
    # Configure and connect to Postgres
    conn = psycopg2.connect(
        host="localhost",
        database="rao_pbems_db",
        user="postgres",
        password="rao123",
        port="5432",
    )
    # Create a cursor
    c = conn.cursor()
    sql_insert_blob_query = """ select * from employees"""

    c.execute(sql_insert_blob_query)

    record = c.fetchall()
    colnames = [desc[0] for desc in c.description]
    # conn.commit()
    #     print("colnames")
    #     print(colnames)
    conn.close()

    return colnames
    # print("Data Inserted Successfully in " + lf + "reports")


def get_report(id):
    # Configure and connect to Postgres
    conn = psycopg2.connect(
        host="localhost",
        database="rao_pbems_db",
        user="postgres",
        password="rao123",
        port="5432",
    )
    # Create a cursor
    c = conn.cursor()
    sql_insert_blob_query = """ select * from daily_reports_table where daily_id = %s """ % id

    c.execute(sql_insert_blob_query)
    record = c.fetchall()
    colnames = [desc[0] for desc in c.description]
    # conn.commit()
    print("record")
    print(record)
    print("Report get complete")
    conn.close()
    return record
    # print("Data Inserted Successfully in " + lf + "reports")


def get_report_header():
    # Configure and connect to Postgres
    conn = psycopg2.connect(
        host="localhost",
        database="rao_pbems_db",
        user="postgres",
        password="rao123",
        port="5432",
    )
    # Create a cursor
    c = conn.cursor()
    sql_insert_blob_query = """ select * from daily_reports_table """

    c.execute(sql_insert_blob_query)
    record = c.fetchall()
    colnames = [desc[0] for desc in c.description]
    # conn.commit()
    #     print("record")
    #     print(record)
    print("Report Header complete")
    #     conn.close()
    return colnames
    # print("Data Inserted Successfully in " + lf + "reports")


def get_employee(id):
    # Configure and connect to Postgres
    conn = psycopg2.connect(
        host="localhost",
        database="rao_pbems_db",
        user="postgres",
        password="rao123",
        port="5432",
    )
    # Create a cursor
    c = conn.cursor()
    sql_insert_blob_query = """ select * from employees where id = %s """ % id

    c.execute(sql_insert_blob_query)
    record = c.fetchall()
    # conn.commit()
    print(f"select * from employees where id = {id}")
    print("Selection Done")
    conn.close()
    return record
    # print("Data Inserted Successfully in " + lf + "reports")


def update_daily_report(id, col, new_time):
    # Configure and connect to Postgres
    conn = psycopg2.connect(
        host="localhost",
        database="rao_pbems_db",
        user="postgres",
        password="rao123",
        port="5432",
    )
    # Create a cursor
    c = conn.cursor()
    #     sql_insert_blob_query = """ select * from daily_reports_table """
    print("Here")
    if type(new_time) == type([1, 2, 3]):
        new_time = lst2pgarr(new_time)
        sql_insert_blob_query = f""" UPDATE daily_reports_table SET {col} = '{new_time}' WHERE daily_id = {id} """
    else:
        sql_insert_blob_query = f""" UPDATE daily_reports_table SET {col} = {new_time} WHERE daily_id = {id} """
    print("sql_insert_blob_query")
    print(sql_insert_blob_query)

    c.execute(sql_insert_blob_query)
    #     record = c.fetchall()
    print("Here 3")

    conn.commit()
    #     print(record)
    conn.close()
    print(f""" UPDATE daily_reports_table SET {col} = {new_time} WHERE daily_id = {id} """)
    print("Updation Done")


#     return record

# print("Data Inserted Successfully in " + lf + "reports")


def zone2id(zone):
    # No facial recognition necessary zones define the id
    zones = ["101", "404", "534"]
    ids = [3, 1, 2]

    index = zones.index(zone)
    return ids[index]
    # return 0

# from tensorflow.keras.models import Sequential, model_from_json

from keras import models
from keras.models import Sequential, model_from_json

# load json and create model
json_file = open("FYPmodel.json", 'r')
loaded_model_json = json_file.read()
json_file.close()
loaded_model = model_from_json(loaded_model_json)
# load weights into new model
loaded_model.load_weights("FYPmodel.h5")

print("Loaded model from disk")
beepy.beep(1)

IMG_SIZE = 224
BATCH_SIZE = 64
EPOCHS = 100

MAX_SEQ_LENGTH = 20
NUM_FEATURES = 2048

global FPS

def build_feature_extractor():
    feature_extractor = keras.applications.InceptionV3(
        weights="imagenet",
        include_top=False,
        pooling="avg",
        input_shape=(IMG_SIZE, IMG_SIZE, 3),
    )
    preprocess_input = keras.applications.inception_v3.preprocess_input

    inputs = keras.Input((IMG_SIZE, IMG_SIZE, 3))
    preprocessed = preprocess_input(inputs)

    outputs = feature_extractor(preprocessed)
    return keras.Model(inputs, outputs, name="feature_extractor")


feature_extractor = build_feature_extractor()


def loaded_prepare_single_video(frames):
    frames = frames[None, ...]
    frame_mask = np.zeros(shape=(1, MAX_SEQ_LENGTH,), dtype="bool")
    frame_features = np.zeros(shape=(1, MAX_SEQ_LENGTH, NUM_FEATURES), dtype="float32")

    for i, batch in enumerate(frames):
        video_length = batch.shape[0]
        length = min(MAX_SEQ_LENGTH, video_length)
        for j in range(length):
            frame_features[i, j, :] = feature_extractor.predict(batch[None, j, :])
        frame_mask[i, :length] = 1  # 1 = not masked, 0 = masked

    return frame_features, frame_mask


def crop_center_square(frame):
    y, x = frame.shape[0:2]
    min_dim = min(y, x)
    start_x = (x // 2) - (min_dim // 2)
    start_y = (y // 2) - (min_dim // 2)
    return frame[start_y : start_y + min_dim, start_x : start_x + min_dim]


def load_video(path, max_frames=0, resize=(IMG_SIZE, IMG_SIZE)):
    cap = cv2.VideoCapture(path)
    frames = []
    try:
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            frame = crop_center_square(frame)
            frame = cv2.resize(frame, resize)
            frame = frame[:, :, [2, 1, 0]]
            frames.append(frame)

            if len(frames) == max_frames:
                break
    finally:
        cap.release()
    return np.array(frames)


def loaded_sequence_prediction(path):
    # class_vocab = label_processor.get_vocabulary()
    class_vocab = ['not_overlock', 'overlock', 'shoot_gun', 'smoke']
    print("class vocab")
    print(class_vocab)

    frames = load_video(os.path.join("test", path))
    frame_features, frame_mask = loaded_prepare_single_video(frames)
    probabilities = loaded_model.predict([frame_features, frame_mask])[0]
    probs = []
    classes = []
    for i in np.argsort(probabilities)[::-1]:
        try:
            print(f"  {class_vocab[i]}: {probabilities[i] * 100} %")
        except:
            print("in except")
        probs.append(float(probabilities[i] * 100))
        classes.append(class_vocab[i])
    return probs, classes


def person_detector(video):
    haar_upper_body_cascade = cv2.CascadeClassifier("haarcascade_upperbody.xml")

    vidcap = cv2.VideoCapture(video)
    ret, first_frame = vidcap.read()

    totalFrames = vidcap.get(cv2.CAP_PROP_FRAME_COUNT)

    vidcap.set(cv2.CAP_PROP_POS_FRAMES, int(totalFrames / 2))
    ret, second_frame = vidcap.read()

    vidcap.set(cv2.CAP_PROP_POS_FRAMES, int(totalFrames - 2))
    ret, last_frame = vidcap.read()

    person_det = False

    frames = [first_frame, second_frame, last_frame]
    print("person_det")

    for frame in frames:
        upper_body = haar_upper_body_cascade.detectMultiScale(
            cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY),
            scaleFactor=1.1,
            minNeighbors=5,
            minSize=(50, 100),
            # Min size for valid detection, changes according to video size or body size in the video.
            flags=cv2.CASCADE_SCALE_IMAGE
        )
        if len(upper_body) != 0:
            person_det = True
        print(person_det)

    return person_det


zone = "101"
# id = zone2id(zone)
id = zone2id(zone)
print(id)
get_report(3)
violation_names = ['smoke', 'shoot_gun']
report_names = ['daily_id', 'working_time', 'working_overtime', 'violations', 'violation_time', 'todays_start_time',
                'todays_end_time', 'absent']

# inferance here

def chart_maker(report):

    print("report in chart maker")
    print(report)

    # Creating dataset
    keys = ['Working Time', 'Not Working Time', 'Break',
            'Violation time']
    wt = report["working_time"]
    vio = report["violation_time"]
    nwt = 200000 - wt - vio
    b = 1

    data = [wt, nwt, 1, vio]

    # Creating explode data
    explode = (0.0, 0.5, 0.0, 0.5)

    # Creating color parameters
    colors = ("seagreen", "cyan", "yellow",
              "red")

    # Wedge properties
    wp = {'linewidth': 1, 'edgecolor': "black"}

    # Creating autocpt arguments
    def func(pct, allvalues):
        absolute = int(pct / 100. * np.sum(allvalues))
        return "{:.1f}%".format(pct, absolute)

    # Creating plot
    fig, ax = plt.subplots(figsize=(10, 7))
    wedges, texts, autotexts = ax.pie(data,
                                      autopct=lambda pct: func(pct, data),
                                      explode=explode,
                                      labels=keys,
                                      shadow=True,
                                      colors=colors,
                                      startangle=90,
                                      wedgeprops=wp,
                                      textprops=dict(color="black"))

    # Adding legend
    ax.legend(wedges, keys,
              title="Key",
              loc="center left",
              bbox_to_anchor=(1, 0, 0.5, 1))

    plt.setp(autotexts, size=8, weight="bold")
    # ax.set_title("Customizing pie chart")
    # plt.figure(figsize=(1.720, 1.280), dpi=100)
    # show plot
    plt.savefig('chart.png')
    # plt.show()


def report_dict_2_excel(report):
    workbook = load_workbook(filename="report_template.xlsx")

    # open workbook
    sheet = workbook.active

    # modify the desired cell
    # sheet["G5"] = str(datetime.date.today().strftime("%d %m %y"))
    global fps
    sheet["G5"].value = str(datetime.date.today())
    sheet["E10"].value = report["name"]
    sheet["E12"].value = report["department"]
    sheet["E8"].value = report["daily_id"]
    sheet["E18"].value = str("{:.2f}".format(float(report["working_time"]) / (60 * fps))) + " Minutes"
    sheet["E20"].value = str("{:.2f}".format(float(report["working_overtime"]) / (60 * fps))) + " Minutes"
    sheet["E22"].value = str("{:.2f}".format(float(report["violation_time"]) / (60 * fps))) + " Minutes"
    sheet["E24"].value = report["policies_applicable"]
    if len(report["violations"]) == 0:
        sheet["C15"].value = "No Violations"
    else:
        col = 'A'
        for i in range(0, len(report['violations'])):
            sheet[col + "15"].value = report["violations"][i]
            col = chr(ord(col) + 2) # col = col + 2

    # Reduce image size to half

    src = cv2.imread('chart.png', cv2.IMREAD_UNCHANGED)
    # percent by which the image is resized
    scale_percent =60
    # calculate the 50 percent of original dimensions
    width = int(src.shape[1] * scale_percent / 100)
    height = int(src.shape[0] * scale_percent / 100)
    # dsize
    dsize = (width, height)
    # resize image
    output = cv2.resize(src, dsize)
    cv2.imwrite('chart.png', output)

    # Add this new chart image to sheet
    chart_img = drawing.image.Image('chart.png')
    chart_img.anchor = 'A27'
    sheet.add_image(chart_img)


    # save the excel file
    filename = "Report " + str(report["daily_id"]) + " for " + str(datetime.date.today().strftime("%d %m"))
    try:
        workbook.save(filename=str('report2excel.xlsx'))
    except:
        o = win32com.client.Dispatch("Excel.Application")
        o.Visible = True
        wb_path = r"C:\Users\Rao Sharjeel\PycharmProjects\FaceNet\report2excel.xlsx"
        wb = o.Workbooks.Open(wb_path)
        wb.Close(True, r"C:\Users\Rao Sharjeel\PycharmProjects\FaceNet\report2excel.xlsx")



    # convert xlsx to pdf
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = False
    wb_path = r"C:\Users\Rao Sharjeel\PycharmProjects\FaceNet\report2excel.xlsx"
    wb = o.Workbooks.Open(wb_path)
    ws_index_list = [1]  # say you want to print these sheets
    path_to_pdf = r"C:\Users\Rao Sharjeel\PycharmProjects\FaceNet\Reports/" + filename + '.pdf'
    ws = wb.Worksheets[0]
    print_area = 'A1:H50'
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesTall = 1
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.PrintArea = print_area
    wb.WorkSheets(ws_index_list).Select()
    wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)
    wb.Close(True, r"C:\Users\Rao Sharjeel\PycharmProjects\FaceNet\report2excel.xlsx")


import cv2
import numpy as np
from collections import deque
import PIL
from tensorflow import keras

import tkinter
import customtkinter
import tkinter.filedialog
from PIL import Image, ImageTk

SEQUENCE_LENGTH = 10
CLASSES_LIST = ['not_overlock', 'overlock', 'smoke']
print("starting now")

model = keras.models.load_model("lrcn/factory_dataset.h5")

def HAR_report_generator(SEQUENCE_LENGTH,video_name):
    vidcap = cv2.VideoCapture(video_name);
    IMAGE_HEIGHT = 250
    IMAGE_WIDTH = 250
    count = 0
    success = True
    predicted_class_name = ''
    percentage = ''
    frames_queue = deque(maxlen=SEQUENCE_LENGTH)

    customtkinter.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
    customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"
    times = []
    total_frames = 0
    # tm1 = time.time()
    while success:
        tm1 = time.time()
        success, image = vidcap.read()
        # image = rotated=cv2.rotate(image, cv2.ROTATE_90_CLOCKWISE)
        if not success:
            break
        cv2.putText(image,
                    f'Prediction: {predicted_class_name}',
                    (50, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 2,
                    (0, 255, 255),
                    2,
                    cv2.LINE_4)
        print('performing', predicted_class_name)
        cv2.imshow('frame', image)
        resized_frame = cv2.resize(image, (250, 250))
        normalized_frame = resized_frame / 255
        frames_queue.append(normalized_frame)
        if len(frames_queue) == SEQUENCE_LENGTH:
            # Pass the normalized frames to the model and get the predicted probabilities.
            predicted_labels_probabilities = model.predict(np.expand_dims(frames_queue, axis=0))[0]
            # print(predicted_labels_probabilities)
            # Get the index of class with highest probability.
            percentage = round(max(predicted_labels_probabilities) * 100, 2)
            predicted_label = np.argmax(predicted_labels_probabilities)
            # Get the class name using the retrieved index.
            predicted_class_name = CLASSES_LIST[predicted_label]
        #             try:
        #                 img_enc = face_recognition.face_encodings(image)[0]
        #                 result1 = face_recognition.compare_faces([img_enc], img_enc2)
        #             except IndexError as e:
        #                 print("Face Not Found")
        #             if result1 == [True]:
        #                 name = 'junaid'
        #             elif result1 == [False]:
        #                 name = 'unknown'

        prediction = predicted_class_name

        id = 19
        # if abs(probs[1] - probs[0]) <= 10:
        #     person_detected = person_detector(test_video)
        #     if prediction == "not_overlock" and not person_detected:
        #         prediction = "not_overlock"
        #
        #     elif prediction == "overlock" and not person_detected:
        #         prediction = "not_overlock"
        #
        #     elif prediction == "overlock" and person_detected:
        #         prediction = "overlock"
        #
        #     elif prediction == "not_overlock" and person_detected:
        #         prediction = "overlock"

        print("prediction")
        print(prediction)
        # prediction = 'smoke'
        # print(prediction)

        # post inference

        current_dr = get_report(id)

        if prediction == "overlock":
            update_daily_report(id, 'working_time',
                                1 + int(current_dr[0][report_names.index('working_time')]))
        elif prediction in violation_names:
            update_daily_report(id, 'violation_time',
                                1 + int(current_dr[0][report_names.index('violation_time')]))
            print("current_dr[0][report_names.index('violations')]")
            print(current_dr[0][report_names.index('violations')])
            #     print(current_dr[0][report_names.index('violations')].append("HELLOOOO"))
            print(type(current_dr[0][report_names.index('violations')]))
            if prediction not in current_dr[0][report_names.index('violations')]:
                check = current_dr[0][report_names.index('violations')]
                print("check")
                print(check)
                check.append(prediction)
                print(check)
                new_list = list(current_dr[0][report_names.index('violations')]).append(prediction)
                #         print("new_list")
                #         print(new_list)
                update_daily_report(id, 'violations', check)
                print("Added to violations")

        count += 1
        print('performing', f'{predicted_class_name} confidence={percentage}%')

        tm2 = time.time()
        t3 = float(tm2 - tm1)
        times.append(t3)
        total_frames += 1
        print("time taken = " + str(float(tm2 - tm1)) + "seconds")

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    vidcap.release()
    # Destroy all the windows
    cv2.destroyAllWindows()

    tm2 = time.time()
    avg_time = 0
    sum = 0
    for i in times:
        sum += i

    fp = total_frames/sum
    global fps
    fps = fp

    print("FPS = ")
    print(fps)

    print(f"sum = {sum} and frames = {total_frames}")

    print("time taken = " + str(float(tm2 - tm1)) + "seconds")

    beepy.beep(1)

    # Report Generator
    # Here work day is done and now it is time to Generate Reports

    emp_details = get_employee(id)[0]
    emp_header = get_employee_header()

    report = get_report(id)[0] + emp_details[:-1]
    header = get_report_header() + emp_header[:-1]

    # employee_dict = dict(zip(emp_header, emp_details))
    report_dict = dict(zip(header, report))
    # print("employee_dict")
    # print(employee_dict)
    print("report_dict")
    print(report_dict)

    # Dict to txt report
    # report_dict = {'report': report_dict}

    # this will write report json to txt file
    # with open("Reports/report " + str(id) + " " + str(datetime.date.today().strftime("%d %m")) + '.txt', 'w+') as file:
    #     file.write(json.dumps(report_dict))  # use `json.loads` to do the reverse

    chart_maker(report_dict)
    report_dict_2_excel(report_dict)

    beepy.beep(4)
    beepy.beep(4)




def HAR_report_generator2():
    test_video = 'F:\FYP\jupyter Video-Classifier-Using-CNN-and-RNN-main\dataset/test/not_overlock/not_overlock_900 (5).mp4'
    print(type(str(test_video)))

    print(test_video)

    tm1 = time.time()

    probs, classes = loaded_sequence_prediction(test_video)

    print("time for inferrance = " + str(time.time() - tm1))
    prediction = classes[0]
    id = 17
    if abs(probs[1] - probs[0]) <= 10:
        person_detected = person_detector(test_video)
        if prediction == "not_overlock" and not person_detected:
            prediction = "not_overlock"

        elif prediction == "overlock" and not person_detected:
            prediction = "not_overlock"

        elif prediction == "overlock" and person_detected:
            prediction = "overlock"

        elif prediction == "not_overlock" and person_detected:
            prediction = "overlock"

    print("prediction")
    print(prediction)
    # prediction = 'smoke'
    # print(prediction)

    # post inferance

    current_dr = get_report(17)

    if prediction == "overlock":
        update_daily_report(id, 'working_time',
                            int(time.time() - tm1) + int(current_dr[0][report_names.index('working_time')]))
    elif prediction in violation_names:
        update_daily_report(id, 'violation_time',
                            int(time.time() - tm1) + int(current_dr[0][report_names.index('violation_time')]))
        print("current_dr[0][report_names.index('violations')]")
        print(current_dr[0][report_names.index('violations')])
        #     print(current_dr[0][report_names.index('violations')].append("HELLOOOO"))
        print(type(current_dr[0][report_names.index('violations')]))
        if prediction not in current_dr[0][report_names.index('violations')]:
            check = current_dr[0][report_names.index('violations')]
            print("check")
            print(check)
            check.append(prediction)
            print(check)
            new_list = list(current_dr[0][report_names.index('violations')]).append(prediction)
            #         print("new_list")
            #         print(new_list)
            update_daily_report(id, 'violations', check)
            print("Added to violations")

    tm2 = time.time()

    print("time taken = " + str(float(tm2 - tm1)) + "seconds")

    beepy.beep(1)

    # Report Generator
    # Here work day is done and now it is time to Generate Reports


    emp_details = get_employee(id)[0]
    emp_header = get_employee_header()

    report = get_report(id)[0] + emp_details[:-1]
    header = get_report_header() + emp_header[:-1]

    # employee_dict = dict(zip(emp_header, emp_details))
    report_dict = dict(zip(header, report))
    # print("employee_dict")
    # print(employee_dict)
    print("report_dict")
    print(report_dict)

    # Dict to txt report
    # report_dict = {'report': report_dict}

    # this will write report json to txt file
    # with open("Reports/report " + str(id) + " " + str(datetime.date.today().strftime("%d %m")) + '.txt', 'w+') as file:
    #     file.write(json.dumps(report_dict))  # use `json.loads` to do the reverse

    chart_maker(report_dict)
    report_dict_2_excel(report_dict)

    beepy.beep(4)
    beepy.beep(4)

    # if prediction == "overlock":
    #     return "working"
    # if prediction == "not_overlock":
    #     return "not Working"


    # return prediction

# print("starting Report generator")
# HAR_report_generator()
# print("Report Generated")


# emp_details = get_employee(3)[0]
# emp_header = get_employee_header()
#
# print(emp_header)
# print(emp_details)
#
# report = get_report(3)[0] + emp_details[:-1]
# header = get_report_header() + emp_header[:-1]
#
# print(report)
# print(header)